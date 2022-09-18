import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import *
from tkinter import filedialog
import tkinter.ttk as ttk
import json

STUDENTS = []
CONFIG = {}

def add_student():

    student = [Entry()]
    student[0].grid(row=len(STUDENTS))
    for lesson in enumerate(CONFIG["lessons"]):
        n, l = lesson
        Label(text=l).grid(row=len(STUDENTS), column=2 + n*2)
        var = StringVar(tk, '2')
        OptionMenu(tk, var, "2", "3", "4", "5").grid(row=len(STUDENTS), column=1 + n * 2)
        student.append(var)
    STUDENTS.append(student)

def save():
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    for l in enumerate(CONFIG["lessons"]):
        num, lesson = l
        sheet.cell(row=1, column=2+num).value = lesson
    sheet.cell(row=1, column=len(CONFIG["lessons"])+2).value = "Среднее"

    for student in enumerate(STUDENTS):
        n, st = student
        sheet.cell(row=2+n, column=1).value = st[0].get()
        for i in range(len(CONFIG["lessons"])):
            sheet.cell(row=2 + n, column=2+i).value = int(st[1+i].get())
        row = 2+n
        column = get_column_letter(len(CONFIG["lessons"])+1)
        sheet.cell(row=row, column=len(CONFIG["lessons"]) + 2).value = "=AVERAGE(B"+str(row)+':'+str(column)+str(row)+')'

    try:
        filename = filedialog.asksaveasfilename(initialdir = "/",title = "Select file",filetypes = (("Excel files","*.xlsx"),("all files","*.*")))
        wb.save(filename)
    except PermissionError:
        print("Ошибка при сохранении")

def load_configuration():
    with open("config.json") as jsfile:
        conf = json.load(jsfile)
    print(conf)
    return conf



if __name__ == '__main__':
    CONFIG = load_configuration()

    tk = Tk()
    tk.geometry("720x480")

    main_menu = Menu(tk)
    tk.config(menu=main_menu)
    main_menu.add_command(label="Добавить ученика", command=add_student)
    main_menu.add_command(label="Сохранить", command=save)


    tk.mainloop()







